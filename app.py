import io
import re
import time
import zipfile
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from xml.etree import ElementTree as ET

# =========================================================
# CONFIG
# =========================================================
APP_TITLE = "VIES + EORI Checker with screenshots"

VIES_REST_URL = "https://ec.europa.eu/taxation_customs/vies/rest-api/check-vat-number"
VIES_SOAP_URL = "https://ec.europa.eu/taxation_customs/vies/services/checkVatService"
EORI_SOAP_URL = "https://ec.europa.eu/taxation_customs/dds2/eos/validation/services/validation"

REQUEST_TIMEOUT = 45
VIES_CONNECT_TIMEOUT = 12

SOAP_ENV_NS = "http://schemas.xmlsoap.org/soap/envelope/"

# Retry VAT / VIES
VIES_MAX_RETRIES = 3
VIES_RETRY_BASE_DELAY = 4  # seconds
VIES_SLOW_COUNTRY_TIMEOUTS = {"FR": 120}
VIES_SLOW_COUNTRY_RETRIES = {"FR": 1}
VIES_SLOW_COUNTRY_BASE_DELAYS = {"FR": 8}
VIES_FR_RECOVERY_WAIT_SECONDS = 45
VIES_FINAL_INPUT_ERRORS = {"INVALID_INPUT_COUNTRY", "INVALID_INPUT_VAT"}

# VAT blacklist interne
BLACKLISTED_VATS = {
    "DE640503057131089",
    "FR09913979589",
    "FR88914081336",
    "IT02562920203",
}

# =========================================================
# PAGE
# =========================================================
st.set_page_config(
    page_title="ATH VAT EORI CHECKER",
    page_icon="logo.png",
    layout="wide"
)

st.sidebar.image("logo.png", width=200)
st.sidebar.markdown("### Athina Logistics")
st.sidebar.caption("Global Access")
st.title(APP_TITLE)
st.caption("Upload invoices Excel -> check VAT (J11) + EORI (J12) -> ZIP captures + recap")

# =========================================================
# HELPERS
# =========================================================
def clean_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def safe_filename(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r'[\\/*?:"<>|]+', "_", text)
    text = text.strip(" .")
    return text or "file"


def split_country_and_vat(raw: str) -> Tuple[str, str]:
    """
    Exemples acceptés :
    PL5263787827
    PL 5263787827
    PL-5263787827
    """
    raw = clean_text(raw).upper()
    raw = raw.replace(" ", "").replace("-", "").replace(".", "").replace("/", "")
    m = re.match(r"^([A-Z]{2})([A-Z0-9+*.]{2,12})$", raw)
    if m:
        return m.group(1), m.group(2)
    return "", raw


def normalize_vat_full(country_code: str, vat_number: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", f"{clean_text(country_code)}{clean_text(vat_number)}".upper())


def normalize_eori(raw: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", clean_text(raw).upper())


def xml_escape(text: str) -> str:
    text = clean_text(text)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def find_text_any_ns(root, tag_name: str) -> str:
    for elem in root.iter():
        if elem.tag.endswith(tag_name):
            return (elem.text or "").strip()
    return ""


# =========================================================
# EXTRACTION INVOICE
# =========================================================
def extract_invoice_data(uploaded_file) -> Dict:
    """
    VAT = J11
    EORI = J12
    Nom entreprise = C11
    Feuille INVOICE si elle existe, sinon feuille active
    """
    file_bytes = uploaded_file.getvalue()
    bio = io.BytesIO(file_bytes)

    try:
        wb = load_workbook(bio, data_only=True)
    except Exception as exc:
        return {
            "ok": False,
            "file": uploaded_file.name,
            "error": f"Impossible d'ouvrir le fichier Excel: {exc}",
        }

    ws = wb["INVOICE"] if "INVOICE" in wb.sheetnames else wb.active

    company_name = clean_text(ws["C11"].value)
    raw_vat = clean_text(ws["J11"].value)
    raw_eori = clean_text(ws["J12"].value)

    country_code, vat_number = split_country_and_vat(raw_vat)
    vat_full = normalize_vat_full(country_code, vat_number)
    eori = normalize_eori(raw_eori)
    is_blacklisted_vat = vat_full in BLACKLISTED_VATS

    errors = []
    if not raw_vat:
        errors.append("J11 vide")
    elif not country_code or not vat_number:
        errors.append(f"J11 invalide: {raw_vat}")

    if not raw_eori:
        errors.append("J12 vide")

    if errors:
        return {
            "ok": False,
            "file": uploaded_file.name,
            "error": " | ".join(errors),
        }

    return {
        "ok": True,
        "file": uploaded_file.name,
        "country_code": country_code,
        "vat_number": vat_number,
        "vat": vat_full,
        "eori": eori,
        "company_name": company_name,
        "is_blacklisted_vat": is_blacklisted_vat,
        "blacklist_alert": "BLACKLIST VAT" if is_blacklisted_vat else "",
        "raw_vat": raw_vat,
        "raw_eori": raw_eori,
    }


# =========================================================
# VIES SOAP
# =========================================================
def build_vies_envelope(country_code: str, vat_number: str) -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope
    xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
    xmlns:urn="urn:ec.europa.eu:taxud:vies:services:checkVat:types">
   <soapenv:Header/>
   <soapenv:Body>
      <urn:checkVat>
         <urn:countryCode>{xml_escape(country_code)}</urn:countryCode>
         <urn:vatNumber>{xml_escape(vat_number)}</urn:vatNumber>
      </urn:checkVat>
   </soapenv:Body>
</soapenv:Envelope>"""


def parse_vies_response(xml_text: str) -> Dict:
    root = ET.fromstring(xml_text)

    fault = root.find(f".//{{{SOAP_ENV_NS}}}Fault")
    if fault is not None:
        faultstring = fault.findtext("faultstring") or "UNKNOWN_ERROR"
        return {"ok": False, "error": faultstring.strip()}

    return {
        "ok": True,
        "country_code": find_text_any_ns(root, "countryCode"),
        "vat_number": find_text_any_ns(root, "vatNumber"),
        "request_date": find_text_any_ns(root, "requestDate"),
        "valid": find_text_any_ns(root, "valid").lower() == "true",
        "name": "" if find_text_any_ns(root, "name") == "---" else find_text_any_ns(root, "name"),
        "address": "" if find_text_any_ns(root, "address") == "---" else find_text_any_ns(root, "address"),
    }


def vies_timeout(country_code: str) -> int:
    country_code = clean_text(country_code).upper()
    return VIES_SLOW_COUNTRY_TIMEOUTS.get(country_code, REQUEST_TIMEOUT)


def vies_request_timeout(country_code: str) -> Tuple[int, int]:
    return (VIES_CONNECT_TIMEOUT, vies_timeout(country_code))


def vies_retry_settings(country_code: str) -> Tuple[int, int]:
    country_code = clean_text(country_code).upper()
    return (
        VIES_SLOW_COUNTRY_RETRIES.get(country_code, VIES_MAX_RETRIES),
        VIES_SLOW_COUNTRY_BASE_DELAYS.get(country_code, VIES_RETRY_BASE_DELAY),
    )


def normalize_vies_error(error: str) -> str:
    error = clean_text(error)
    if not error:
        return "UNKNOWN_ERROR"
    return error.split(":", 1)[0].strip().upper()


def parse_vies_rest_error(data: Dict) -> str:
    wrappers = data.get("errorWrappers")
    if isinstance(wrappers, list) and wrappers:
        parts = []
        for wrapper in wrappers:
            if not isinstance(wrapper, dict):
                continue
            code = clean_text(wrapper.get("error"))
            message = clean_text(wrapper.get("message"))
            if code and message:
                parts.append(f"{code}: {message}")
            elif code:
                parts.append(code)
            elif message:
                parts.append(message)
        if parts:
            return " | ".join(parts)

    message = clean_text(data.get("message"))
    if message:
        return message

    return "REST_INVALID_RESPONSE"


def parse_vies_rest_response(data: Dict) -> Dict:
    if data.get("actionSucceed") is False or "errorWrappers" in data:
        return {"ok": False, "error": parse_vies_rest_error(data), "method": "REST"}

    if "valid" not in data:
        return {"ok": False, "error": "REST_INVALID_RESPONSE", "method": "REST"}

    name = clean_text(data.get("name") or data.get("traderName"))
    address = clean_text(data.get("address"))

    return {
        "ok": True,
        "country_code": clean_text(data.get("countryCode")),
        "vat_number": clean_text(data.get("vatNumber")),
        "request_date": clean_text(data.get("requestDate")),
        "valid": data.get("valid") is True,
        "name": "" if name == "---" else name,
        "address": "" if address == "---" else address,
        "method": "REST",
    }


def check_vat_rest_once(country_code: str, vat_number: str) -> Dict:
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": "VAT-EORI-Checker/1.0",
    }
    payload = {
        "countryCode": country_code,
        "vatNumber": vat_number,
    }

    try:
        response = requests.post(
            VIES_REST_URL,
            json=payload,
            headers=headers,
            timeout=vies_request_timeout(country_code),
        )

        try:
            data = response.json()
        except ValueError:
            if response.ok:
                return {"ok": False, "error": "INVALID_JSON_RESPONSE", "method": "REST"}
            return {"ok": False, "error": f"HTTP_{response.status_code}", "method": "REST"}

        if not response.ok:
            error = parse_vies_rest_error(data)
            return {"ok": False, "error": error or f"HTTP_{response.status_code}", "method": "REST"}

        return parse_vies_rest_response(data)

    except requests.Timeout:
        return {"ok": False, "error": "TIMEOUT", "method": "REST"}
    except requests.RequestException as exc:
        return {"ok": False, "error": f"HTTP_ERROR: {exc}", "method": "REST"}
    except Exception as exc:
        return {"ok": False, "error": f"REST_UNEXPECTED_ERROR: {exc}", "method": "REST"}


def check_vat_soap_once(country_code: str, vat_number: str) -> Dict:
    headers = {
        "Content-Type": "text/xml; charset=utf-8",
        "SOAPAction": '""',
        "User-Agent": "VAT-EORI-Checker/1.0",
    }

    try:
        response = requests.post(
            VIES_SOAP_URL,
            data=build_vies_envelope(country_code, vat_number).encode("utf-8"),
            headers=headers,
            timeout=vies_request_timeout(country_code),
        )
        response.raise_for_status()
        result = parse_vies_response(response.text)
        result["method"] = "SOAP"
        return result

    except requests.Timeout:
        return {"ok": False, "error": "TIMEOUT", "method": "SOAP"}
    except requests.RequestException as exc:
        return {"ok": False, "error": f"HTTP_ERROR: {exc}", "method": "SOAP"}
    except ET.ParseError:
        return {"ok": False, "error": "INVALID_XML_RESPONSE", "method": "SOAP"}
    except Exception as exc:
        return {"ok": False, "error": f"UNEXPECTED_ERROR: {exc}", "method": "SOAP"}


def check_vat_once(country_code: str, vat_number: str) -> Dict:
    if not re.fullmatch(r"[A-Z]{2}", country_code or ""):
        return {"ok": False, "error": "INVALID_INPUT_COUNTRY"}

    if not re.fullmatch(r"[0-9A-Za-z+*.]{2,12}", vat_number or ""):
        return {"ok": False, "error": "INVALID_INPUT_VAT"}

    rest_result = check_vat_rest_once(country_code, vat_number)
    if rest_result.get("ok"):
        return rest_result

    rest_error_code = normalize_vies_error(rest_result.get("error"))
    if rest_error_code in VIES_FINAL_INPUT_ERRORS:
        return rest_result

    soap_result = check_vat_soap_once(country_code, vat_number)
    if soap_result.get("ok"):
        return soap_result

    return {
        "ok": False,
        "error": f"REST {rest_result.get('error', 'UNKNOWN')} | SOAP {soap_result.get('error', 'UNKNOWN')}",
        "method": "REST+SOAP",
    }


def check_vat_with_retry(country_code: str, vat_number: str, max_retries: int = None, status_callback=None) -> Dict:
    """
    Règle :
    - VALID   -> stop
    - INVALID -> stop
    - ERROR   -> retry
    """
    default_retries, base_delay = vies_retry_settings(country_code)
    max_retries = max_retries or default_retries
    last_result = None

    for attempt in range(1, max_retries + 1):
        if status_callback:
            status_callback("checking", attempt, max_retries, "", 0)

        result = check_vat_once(country_code, vat_number)
        result["attempts"] = attempt
        last_result = result

        # Si réponse propre de VIES -> on stop directement
        if result.get("ok"):
            return result

        # Si input invalide -> inutile de retry
        err = normalize_vies_error(result.get("error"))
        if err in VIES_FINAL_INPUT_ERRORS:
            return result

        # Sinon erreur VIES/réseau/XML -> retry
        if attempt < max_retries:
            sleep_time = base_delay * attempt
            if status_callback:
                status_callback("waiting", attempt, max_retries, err, sleep_time)
            time.sleep(sleep_time)

    return last_result or {"ok": False, "error": "NO_RESPONSE", "attempts": 0}


def is_final_vat_result(result: Dict) -> bool:
    if result.get("ok"):
        return True
    return normalize_vies_error(result.get("error")) in VIES_FINAL_INPUT_ERRORS


def is_retryable_vat_error(result: Dict) -> bool:
    return not is_final_vat_result(result)


def should_cache_vat_result(result: Dict) -> bool:
    return is_final_vat_result(result)


def vat_key_from_source(row: Dict) -> Tuple[str, str]:
    return row["country_code"], row["vat_number"]


def apply_vat_result_to_matching_rows(results: List[Dict], vat_key: Tuple[str, str], vat_result: Dict) -> int:
    updated = 0
    for item in results:
        source = item["source"]
        if vat_key_from_source(source) != vat_key:
            continue

        result_copy = dict(vat_result)
        if updated:
            result_copy["from_cache"] = True
        item["vat_result"] = result_copy
        item["vat_attempts"] = result_copy.get("attempts", item.get("vat_attempts", 1))
        updated += 1
    return updated


def retry_french_vat_errors(results: List[Dict], vat_cache: Dict, status_box) -> int:
    candidates = []
    seen = set()

    for item in results:
        source = item["source"]
        vat_result = item["vat_result"]
        if source.get("country_code") != "FR":
            continue
        if not is_retryable_vat_error(vat_result):
            continue

        vat_key = vat_key_from_source(source)
        if vat_key in seen:
            continue
        seen.add(vat_key)
        candidates.append((vat_key, source))

    if not candidates:
        return 0

    status_box.write(
        f"French VAT temporary errors detected. Waiting {VIES_FR_RECOVERY_WAIT_SECONDS}s before retry."
    )
    time.sleep(VIES_FR_RECOVERY_WAIT_SECONDS)

    recovered = 0
    total = len(candidates)

    for index, (vat_key, source) in enumerate(candidates, start=1):
        status_box.write(
            f"French VAT retry {index}/{total}: {source['file']} -> {source['vat']}"
        )
        retry_result = check_vat_with_retry(
            source["country_code"],
            source["vat_number"],
            max_retries=1,
        )

        if should_cache_vat_result(retry_result):
            vat_cache[vat_key] = dict(retry_result)

        updated = apply_vat_result_to_matching_rows(results, vat_key, retry_result)
        if retry_result.get("ok"):
            recovered += updated

    return recovered


# =========================================================
# EORI SOAP
# =========================================================
def build_eori_envelope(eori: str) -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope
    xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
    xmlns:eori="http://eori.ws.eos.dds.s/">
   <soapenv:Header/>
   <soapenv:Body>
      <eori:validateEORI>
         <eori:eori>{xml_escape(eori)}</eori:eori>
      </eori:validateEORI>
   </soapenv:Body>
</soapenv:Envelope>"""


def parse_eori_response(xml_text: str) -> Dict:
    root = ET.fromstring(xml_text)

    fault = root.find(f".//{{{SOAP_ENV_NS}}}Fault")
    if fault is not None:
        faultstring = fault.findtext("faultstring") or "UNKNOWN_ERROR"
        return {"ok": False, "error": faultstring.strip()}

    request_date = find_text_any_ns(root, "requestDate")
    error_description = find_text_any_ns(root, "errorDescription")

    results = []
    for elem in root.iter():
        if elem.tag.endswith("result"):
            result_data = {
                "eori": "",
                "status": "",
                "statusDescr": "",
                "errorReason": "",
                "name": "",
                "address": "",
                "street": "",
                "postalCode": "",
                "city": "",
                "country": "",
            }
            for child in elem:
                tag = child.tag.split("}")[-1]
                result_data[tag] = (child.text or "").strip()
            if result_data["eori"] or result_data["statusDescr"] or result_data["status"]:
                results.append(result_data)

    first = results[0] if results else {}

    status_descr = clean_text(first.get("statusDescr", ""))
    status_descr_l = status_descr.lower()

    is_valid = False
    if status_descr_l:
        if "invalid" not in status_descr_l and "valid" in status_descr_l:
            is_valid = True

    return {
        "ok": True,
        "request_date": request_date,
        "error_description": error_description,
        "eori": clean_text(first.get("eori", "")),
        "status": clean_text(first.get("status", "")),
        "status_descr": status_descr,
        "error_reason": clean_text(first.get("errorReason", "")),
        "name": clean_text(first.get("name", "")),
        "address": clean_text(first.get("address", "")),
        "street": clean_text(first.get("street", "")),
        "postal_code": clean_text(first.get("postalCode", "")),
        "city": clean_text(first.get("city", "")),
        "country": clean_text(first.get("country", "")),
        "valid": is_valid,
    }


def check_eori(eori: str) -> Dict:
    if not eori:
        return {"ok": False, "error": "EMPTY_EORI"}

    headers = {
        "Content-Type": "text/xml; charset=utf-8",
        "SOAPAction": '""',
        "User-Agent": "VAT-EORI-Checker/1.0",
    }

    try:
        response = requests.post(
            EORI_SOAP_URL,
            data=build_eori_envelope(eori).encode("utf-8"),
            headers=headers,
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        return parse_eori_response(response.text)

    except requests.Timeout:
        return {"ok": False, "error": "TIMEOUT"}
    except requests.RequestException as exc:
        return {"ok": False, "error": f"HTTP_ERROR: {exc}"}
    except ET.ParseError:
        return {"ok": False, "error": "INVALID_XML_RESPONSE"}
    except Exception as exc:
        return {"ok": False, "error": f"UNEXPECTED_ERROR: {exc}"}


# =========================================================
# FONTS / DRAW
# =========================================================
def get_font(size: int, bold: bool = False):
    candidates = []
    if bold:
        candidates = [
            "arialbd.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        ]
    else:
        candidates = [
            "arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        ]

    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass

    return ImageFont.load_default()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font, max_width: int) -> List[str]:
    text = clean_text(text)
    if not text:
        return [""]

    words = text.replace("\r", "").replace("\n", " ").split()
    lines = []
    current = ""

    for word in words:
        test = word if not current else current + " " + word
        bbox = draw.textbbox((0, 0), test, font=font)
        width = bbox[2] - bbox[0]
        if width <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word

    if current:
        lines.append(current)

    return lines if lines else [""]


# =========================================================
# IMAGE VAT
# =========================================================
def render_vat_image(vat_result: Dict, source_row: Dict) -> bytes:
    width, height = 1035, 570
    img = Image.new("RGB", (width, height), (250, 250, 250))
    draw = ImageDraw.Draw(img)

    title_font = get_font(30, bold=True)
    text_font = get_font(15)
    label_font = get_font(15, bold=True)
    value_font = get_font(14)
    status_font = get_font(18, bold=True)
    button_font = get_font(14, bold=True)

    draw.text((28, 12), "VIES VAT number validation", font=title_font, fill=(45, 45, 45))

    intro = (
        "You can verify the validity of a VAT number issued by any Member State / Northern Ireland "
        "by selecting that Member State / Northern Ireland from the drop-down menu provided, and "
        "entering the number to be validated."
    )
    intro_lines = wrap_text(draw, intro, text_font, 980)
    y_intro = 84
    for line in intro_lines:
        draw.text((28, y_intro), line, font=text_font, fill=(15, 23, 42))
        y_intro += 26

    box = (28, 159, width - 40, height - 16)
    draw.rectangle(box, fill=(250, 250, 250), outline=(75, 75, 75), width=1)

    is_valid = vat_result.get("ok") and vat_result.get("valid") is True
    if vat_result.get("ok"):
        status_text = "Yes, valid VAT number" if is_valid else "No, invalid VAT number"
    else:
        status_text = "VAT validation error"
    status_color = (55, 120, 28) if is_valid else (170, 40, 40)

    icon_x, icon_y = 68, 199
    draw.ellipse((icon_x - 8, icon_y - 8, icon_x + 8, icon_y + 8), fill=status_color)
    draw.line((icon_x - 5, icon_y, icon_x - 1, icon_y + 5, icon_x + 7, icon_y - 5), fill=(255, 255, 255), width=3)
    draw.text((91, 190), status_text, font=status_font, fill=status_color)

    labels = (
        "Member State / Northern Ireland",
        "VAT Number",
        "Date when request received",
        "Name",
        "Address",
        "Consultation Number",
    )

    if vat_result.get("ok"):
        request_date = vat_result.get("request_date", "")
        if request_date:
            try:
                dt = datetime.fromisoformat(request_date)
                if dt.hour or dt.minute or dt.second:
                    request_date = dt.strftime("%d/%m/%Y %H:%M:%S")
                else:
                    request_date = f"{dt.strftime('%d/%m/%Y')} {datetime.now().strftime('%H:%M:%S')}"
            except Exception:
                pass

        values = [
            vat_result.get("country_code", source_row.get("country_code", "")),
            vat_result.get("vat_number", source_row.get("vat_number", "")),
            request_date,
            vat_result.get("name", "") or "---",
            vat_result.get("address", "") or "---",
            "",
        ]
    else:
        values = [
            source_row.get("country_code", ""),
            source_row.get("vat_number", ""),
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            f"Error: {vat_result.get('error', 'UNKNOWN')}",
            "---",
            "",
        ]

    left_x = 72
    value_x = 633
    line_x1 = 58
    line_x2 = width - 40
    start_y = 242
    row_gap = 41

    for i, (lab, val) in enumerate(zip(labels, values)):
        y = start_y + i * row_gap
        draw.line((line_x1, y + 24, line_x2, y + 24), fill=(210, 210, 210), width=1)
        draw.text((left_x, y), lab, font=label_font, fill=(30, 30, 30))

        wrapped = wrap_text(draw, str(val), value_font, line_x2 - value_x - 10)
        yy = y
        for line in wrapped[:2]:
            draw.text((value_x, yy), line, font=value_font, fill=(30, 30, 30))
            yy += 18

    printer_x, printer_y = width - 57, 193
    draw.rectangle((printer_x, printer_y, printer_x + 14, printer_y + 10), outline=(0, 75, 160), width=2)
    draw.rectangle((printer_x + 3, printer_y + 8, printer_x + 17, printer_y + 14), outline=(0, 75, 160), width=2)
    draw.rectangle((printer_x + 4, printer_y - 5, printer_x + 13, printer_y), outline=(0, 75, 160), width=2)

    draw.rectangle((934, 487, 995, 526), fill=(255, 255, 255), outline=(0, 75, 160), width=1)
    draw.text((951, 499), "Back", font=button_font, fill=(0, 75, 160))

    output = io.BytesIO()
    img.save(output, format="PNG")
    return output.getvalue()


# =========================================================
# IMAGE EORI
# =========================================================
def render_eori_image(eori_result: Dict, source_row: Dict) -> bytes:
    width, height = 1052, 684
    img = Image.new("RGB", (width, height), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    title_font = get_font(29, bold=True)
    section_font = get_font(20, bold=True)
    text_font = get_font(14)
    label_font = get_font(15, bold=True)
    value_font = get_font(14)
    button_font = get_font(14, bold=True)

    draw.text((28, 15), "Validation de numéro EORI", font=title_font, fill=(45, 45, 45))

    import_date = datetime.now().strftime("%d-%m-%Y")
    draw.text((28, 63), f"Dernière date d'importation : {import_date}", font=text_font, fill=(0, 0, 0))

    draw.rectangle((28, 120, width - 27, 163), fill=(7, 80, 160))
    draw.text((43, 132), "Retrieve EORI number validation", font=section_font, fill=(255, 255, 255))

    draw.rectangle((28, 163, width - 27, 282), fill=(245, 245, 245))
    draw.text(
        (43, 183),
        "Lancez une demande de validation en saisissant le numéro EORI et en cliquant sur \"Valider\".",
        font=text_font,
        fill=(0, 0, 0),
    )

    draw.text((43, 220), "numéro EORI", font=label_font, fill=(40, 40, 40))
    draw.rectangle((236, 207, 816, 246), fill=(255, 255, 255), outline=(110, 110, 110), width=1)
    draw.text((252, 219), source_row.get("eori", ""), font=value_font, fill=(35, 35, 35))
    draw.rectangle((816, 207, 893, 246), fill=(7, 80, 160))
    draw.text((831, 219), "Valider", font=button_font, fill=(255, 255, 255))

    start_y = 342
    left_x = 42
    value_x = 437
    line_x1 = 28
    line_x2 = width - 27
    row_gap = 44

    if eori_result.get("ok"):
        request_date = eori_result.get("request_date", "")
        if request_date:
            try:
                dt = datetime.fromisoformat(request_date)
                request_date = dt.strftime("%d/%m/%Y")
            except Exception:
                request_date = request_date[:10] if len(request_date) >= 10 else request_date

        status_text = eori_result.get("status_descr", "") or (
            "This EORI number is valid." if eori_result.get("valid") else "Invalid / error"
        )
        country = eori_result.get("country", "")
        if clean_text(country).upper() in {"CN", "CHINA"}:
            country = "Chine"

        rows = [
            ("Date de la demande:", request_date),
            ("", status_text),
            ("Nom", eori_result.get("name", "")),
            ("Adresse", eori_result.get("address", "")),
            ("Street number", eori_result.get("street", "")),
            ("Postal code", eori_result.get("postal_code", "")),
            ("Ville", eori_result.get("city", "")),
            ("Pays:", country),
        ]
    else:
        rows = [
            ("Date de la demande:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("", f"Error: {eori_result.get('error', 'UNKNOWN')}"),
            ("Nom", ""),
            ("Adresse", ""),
            ("Street number", ""),
            ("Postal code", ""),
            ("Ville", ""),
            ("Pays:", ""),
        ]

    for i, (lab, val) in enumerate(rows):
        y = start_y + i * row_gap
        draw.line((line_x1, y + 27, line_x2, y + 27), fill=(205, 205, 205), width=1)
        if not lab:
            draw.text((left_x, y), clean_text(val), font=label_font, fill=(45, 45, 45))
            continue

        if lab:
            draw.text((left_x, y), lab, font=label_font, fill=(45, 45, 45))
        wrapped = wrap_text(draw, clean_text(val), value_font, line_x2 - value_x - 15)
        yy = y
        for line in wrapped[:2]:
            draw.text((value_x, yy), line, font=value_font, fill=(45, 45, 45))
            yy += 18

    output = io.BytesIO()
    img.save(output, format="PNG")
    return output.getvalue()


# =========================================================
# DATAFRAME / ZIP
# =========================================================
def results_to_dataframe(results: List[Dict]) -> pd.DataFrame:
    rows = []

    for item in results:
        source = item["source"]
        vat_result = item["vat_result"]
        eori_result = item["eori_result"]
        vat_attempts = item.get("vat_attempts", 1)

        vat_status = "TEMP ERROR"
        if vat_result.get("ok"):
            vat_status = "VALID" if vat_result.get("valid") else "INVALID"
        elif normalize_vies_error(vat_result.get("error")) in VIES_FINAL_INPUT_ERRORS:
            vat_status = "ERROR"

        eori_status = "ERROR"
        if eori_result.get("ok"):
            eori_status = "VALID" if eori_result.get("valid") else "INVALID"

        rows.append({
            "file": source.get("file", ""),
            "company_name": source.get("company_name", ""),
            "vat": source.get("vat", ""),
            "country_code": source.get("country_code", ""),
            "vat_number": source.get("vat_number", ""),
            "blacklist_alert": source.get("blacklist_alert", ""),
            "vat_status": vat_status,
            "vat_error": "" if vat_result.get("ok") else vat_result.get("error", ""),
            "vat_attempts": vat_attempts,
            "vat_method": vat_result.get("method", ""),
            "vat_from_cache": "YES" if vat_result.get("from_cache") else "",
            "vat_request_date": vat_result.get("request_date", ""),
            "vat_name": vat_result.get("name", ""),
            "vat_address": vat_result.get("address", ""),
            "eori": source.get("eori", ""),
            "eori_status": eori_status,
            "eori_error": "" if eori_result.get("ok") else eori_result.get("error", ""),
            "eori_from_cache": "YES" if eori_result.get("from_cache") else "",
            "eori_request_date": eori_result.get("request_date", ""),
            "eori_status_descr": eori_result.get("status_descr", ""),
            "eori_name": eori_result.get("name", ""),
            "eori_address": eori_result.get("address", ""),
            "eori_street": eori_result.get("street", ""),
            "eori_postal_code": eori_result.get("postal_code", ""),
            "eori_city": eori_result.get("city", ""),
            "eori_country": eori_result.get("country", ""),
        })

    return pd.DataFrame(rows)


def display_results_dataframe(summary_df: pd.DataFrame):
    columns = [
        "file",
        "company_name",
        "vat",
        "vat_status",
        "vat_error",
        "vat_attempts",
        "vat_method",
        "vat_from_cache",
        "eori",
        "eori_status",
        "eori_error",
    ]
    visible_columns = [col for col in columns if col in summary_df.columns]
    st.dataframe(summary_df[visible_columns], use_container_width=True)

    with st.expander("Full result details"):
        st.dataframe(summary_df, use_container_width=True)


def make_zip(vat_images: List[Tuple[str, bytes]], eori_images: List[Tuple[str, bytes]], summary_df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for filename, img_bytes in vat_images:
            zf.writestr(f"captures_vat/{filename}", img_bytes)

        for filename, img_bytes in eori_images:
            zf.writestr(f"captures_eori/{filename}", img_bytes)

        excel_buf = io.BytesIO()
        with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name="VAT_EORI_RESULTS")
        zf.writestr("vat_eori_results.xlsx", excel_buf.getvalue())

    buf.seek(0)
    return buf.getvalue()


# =========================================================
# UI
# =========================================================
st.markdown(
    """
**Mode invoices directes**
- upload one or more `.xlsx` files
- launche verification
- output = ZIP with VAT captures + EORI + Excel summary
"""
)

uploaded_files = st.file_uploader(
    "Upload your invoices Excel",
    type=["xlsx"],
    accept_multiple_files=True,
)

if uploaded_files:
    parsed_rows = []
    extract_errors = []

    for f in uploaded_files:
        extracted = extract_invoice_data(f)
        if extracted["ok"]:
            parsed_rows.append(extracted)
        else:
            extract_errors.append(extracted)

    st.subheader("Aperçu extraction")
    if parsed_rows:
        preview_df = pd.DataFrame(
            [
                {
                    "nom de facture": row["file"],
                    "VAT": row["vat"],
                    "Eori": row["eori"],
                    "nom d'entreprise": row["company_name"],
                    "alerte blacklist": row["blacklist_alert"],
                }
                for row in parsed_rows
            ]
        )
        st.dataframe(preview_df, use_container_width=True)

        blacklisted_rows = [row for row in parsed_rows if row.get("is_blacklisted_vat")]
        if blacklisted_rows:
            st.error(
                "⚠️ VAT blacklist détecté : "
                + ", ".join(f"{row['file']} -> {row['vat']}" for row in blacklisted_rows)
            )
    else:
        st.warning("Aucune invoice exploitable trouvée.")

    if extract_errors:
        st.subheader("Fichiers en erreur")
        st.dataframe(pd.DataFrame(extract_errors), use_container_width=True)

    if parsed_rows and st.button("Launch verification VAT + EORI", type="primary"):
        results = []
        vat_images = []
        eori_images = []
        vat_cache = {}
        eori_cache = {}

        progress = st.progress(0)
        status_box = st.empty()

        total = len(parsed_rows)

        for i, row in enumerate(parsed_rows, start=1):
            status_box.write(
                f"Traitement {i}/{total} : {row['file']} -> VAT {row['vat']} | EORI {row['eori']}"
            )

            def update_vat_status(phase, attempt, max_retries, error, wait_seconds):
                if phase == "checking":
                    status_box.write(
                        f"Traitement {i}/{total} : {row['file']} -> VAT attempt {attempt}/{max_retries} | EORI {row['eori']}"
                    )
                else:
                    status_box.write(
                        f"Traitement {i}/{total} : {row['file']} -> VAT retry {attempt}/{max_retries} "
                        f"after {wait_seconds}s ({error}) | EORI {row['eori']}"
                    )

            vat_key = (row["country_code"], row["vat_number"])
            if vat_key in vat_cache:
                vat_result = dict(vat_cache[vat_key])
                vat_result["from_cache"] = True
                status_box.write(
                    f"Traitement {i}/{total} : {row['file']} -> VAT reused from cache | EORI {row['eori']}"
                )
            else:
                vat_result = check_vat_with_retry(
                    row["country_code"],
                    row["vat_number"],
                    status_callback=update_vat_status,
                )
                if should_cache_vat_result(vat_result):
                    vat_cache[vat_key] = dict(vat_result)

            vat_attempts = vat_result.get("attempts", 1)

            eori_key = row["eori"]
            if eori_key in eori_cache:
                eori_result = dict(eori_cache[eori_key])
                eori_result["from_cache"] = True
            else:
                eori_result = check_eori(row["eori"])
                if eori_result.get("ok"):
                    eori_cache[eori_key] = dict(eori_result)

            results.append({
                "source": row,
                "vat_result": vat_result,
                "vat_attempts": vat_attempts,
                "eori_result": eori_result,
            })

            vat_img = render_vat_image(vat_result, row)
            eori_img = render_eori_image(eori_result, row)

            base_name = safe_filename(row["file"].rsplit(".", 1)[0])

            vat_images.append((
                f"{base_name}_VAT.png",
                vat_img,
            ))

            eori_images.append((
                f"{base_name}_EORI.png",
                eori_img,
            ))

            progress.progress(i / total)

        recovered_fr = retry_french_vat_errors(results, vat_cache, status_box)
        if recovered_fr:
            status_box.success(f"French VAT retry recovered {recovered_fr} result(s).")
            vat_images = []
            for item in results:
                row = item["source"]
                vat_img = render_vat_image(item["vat_result"], row)
                base_name = safe_filename(row["file"].rsplit(".", 1)[0])
                vat_images.append((f"{base_name}_VAT.png", vat_img))

        status_box.success("Traitement terminé.")

        summary_df = results_to_dataframe(results)

        st.subheader("Résultats")
        display_results_dataframe(summary_df)

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("VAT VALID", int((summary_df["vat_status"] == "VALID").sum()))
        c2.metric("VAT INVALID", int((summary_df["vat_status"] == "INVALID").sum()))
        c3.metric("VAT TEMP ERROR", int((summary_df["vat_status"] == "TEMP ERROR").sum()))
        c4.metric("EORI VALID", int((summary_df["eori_status"] == "VALID").sum()))
        c5.metric("EORI ERROR/INVALID", int((summary_df["eori_status"] != "VALID").sum()))

        zip_bytes = make_zip(vat_images, eori_images, summary_df)

        st.download_button(
            label="Télécharger le ZIP (captures VAT + EORI + recap Excel)",
            data=zip_bytes,
            file_name="vat_eori_results.zip",
            mime="application/zip",
        )
